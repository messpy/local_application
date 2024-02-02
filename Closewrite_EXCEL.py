import openpyxl

# Excelファイルのパス
excel_file_path = 'test.xlsx'

# ワークブックを開く
workbook = openpyxl.load_workbook(excel_file_path)

rowclm = 5

def first_choice():   
    # シートを選択
    sheet_names = workbook.sheetnames
    i = 0
    for shtname in sheet_names:
        print(str(i)+shtname)
        i += 1
        shtnum=int(input("何番目のシートを表示しますか？"))
        center_cell_address = input("Cellを選択してください (例: E6): ")  # 修正
        sht = workbook.worksheets[shtnum]
        center_cell = sht[center_cell_address]  # 修正
        return sht, center_cell

def cell_display(rowclm, sht, center_cell):        
    center_row, center_column = center_cell.row, center_cell.column

    # ヘッダー行
    print("   ", end="")
    for col_offset in range(-2, 3):
        target_column = center_column + col_offset
        if 1 <= target_column <= sht.max_column:
            print(f"{sht.cell(row=1, column=target_column).column_letter:^10}", end="")
        else:
            print(" " * 10, end="")
    print()  # 改行

    for row_offset in range(-2, 3):
        target_row = center_row + row_offset

        if 1 <= target_row <= sht.max_row:
            print(f"{target_row:2d} ", end="")
            for col_offset in range(-2, 3):
                target_column = center_column + col_offset
                
                if 1 <= target_column <= sht.max_column:
                    current_cell = sht.cell(row=target_row, column=target_column)
                    print(f"{str(current_cell.value):^10}", end="")
                else:
                    print(" " * 10, end="")  # 列が存在しない場合は空白セル

            print()  # 改行

if __name__ == "__main__":
    sht, center_cell = first_choice()
    cell_display(rowclm, sht, center_cell)

    while True:
        # キーボード入力を待つ
        user_input = input("操作を選択してください (W: 上, A: 左, S: 下, D: 右, R: 書き込み, Q: 終了): ")
        
        if user_input.lower() == 'q':
            break  # Qが押されたら終了
        elif user_input.lower() == 'r':
            # Rが押されたら書き込み処理
            new_value = input("書き込む値を入力してください: ")
            left_cell_address = sht.cell(row=center_cell.row, column=center_cell.column - 1).coordinate
            sht[left_cell_address].value = new_value
            print("書き込みが完了しました。")
            
            # ワークブックを保存
            workbook.save(excel_file_path)
            
            # 書き込み後にセル表示を更新
            sht, center_cell = first_choice()
            cell_display(rowclm, sht, center_cell)
        else:
            # 移動方向に応じてセル表示を更新
            if user_input.lower() == 'w':
                center_cell_address = sht.cell(row=center_cell.row - 1, column=center_cell.column).coordinate
            elif user_input.lower() == 'a':
                center_cell_address = sht.cell(row=center_cell.row, column=center_cell.column - 1).coordinate
            elif user_input.lower() == 's':
                center_cell_address = sht.cell(row=center_cell.row + 1, column=center_cell.column).coordinate
            elif user_input.lower() == 'd':
                center_cell_address = sht.cell(row=center_cell.row, column=center_cell.column + 1).coordinate
            else:
                print("無効な操作です。再入力してください。")
                continue

            center_cell = sht[center_cell_address]
            cell_display(rowclm, sht, center_cell)

    # ワークブックを保存
    workbook.save(excel_file_path)
    # ワークブックを閉じる
    workbook.close()
